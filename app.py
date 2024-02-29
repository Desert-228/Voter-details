from flask import Flask, render_template, request, send_file
from concurrent.futures import ThreadPoolExecutor
from pdf2image import convert_from_path
import openpyxl
from PIL import Image
import pytesseract
import re
import os
import pyocr 
import pyocr.builders
tools = pyocr.get_available_tools() 
ocr_tool = tools[0]

app = Flask(__name__)
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

def correct_gender_spelling(gender):
    if gender and gender.lower().startswith('m'):
        return 'Male'
    elif gender and gender.lower().startswith('f'):
        return 'Female'
    else:
        return gender

def process_image(im, idx, ws):
    basewidth = 1700
    wpercent = (basewidth / float(im.size[0]))
    hsize = int((float(im.size[1]) * float(wpercent)))
    im = im.resize((basewidth, hsize), Image.LANCZOS)
    coordinates =  [
        (39, 79, 435, 294), (584, 79, 980, 294), (1129, 79, 1525, 294),
        (39, 308, 435, 523), (584, 308, 980, 523), (1129, 308, 1525, 523),
        (39, 537, 435, 752), (584, 537, 980, 752), (1129, 537, 1525, 752),
        (39, 766, 435, 981), (584, 766, 980, 981), (1129, 766, 1525, 981),
        (39, 995, 435, 1210), (584, 995, 980, 1210), (1129, 995, 1525, 1210),
        (39, 1224, 435, 1439), (584, 1224, 980, 1439), (1129, 1224, 1525, 1439),
        (39, 1453, 435, 1668), (584, 1453, 980, 1668), (1129, 1453, 1525, 1668),
        (39, 1682, 435, 1897), (584, 1682, 980, 1897), (1129, 1682, 1525, 1897),
        (39, 1911, 435, 2126), (584, 1911, 980, 2126), (1129, 1911, 1525, 2126),
        (39, 2140, 435, 2355), (584, 2140, 980, 2355), (1129, 2140,1525,2355)
        ]
    coordinates1 =  [
    (370, 79, 570, 123), (900, 79, 1110, 123), (1450, 79, 1660, 123),
    (370,308,570,354), (900,308,1110,354), (1450,308,1660,354),
    (370, 537, 570, 582), (900, 537, 1110, 582), (1450, 537, 1660, 582),
    (370, 760, 570, 811), (900, 760, 1110, 811), (1450, 760, 1660, 811),
    (370, 990, 570, 1035), (900, 990, 1110, 1035), (1450, 990, 1660, 1035),
    (370, 1220, 570, 1265), (900, 1220, 1110, 1265), (1450, 1220, 1660, 1265),
    (370, 1446, 570, 1490), (900, 1446, 1110, 1490), (1450, 1446, 1660, 1490),
    (370, 1672, 570, 1718), (900, 1672, 1110, 1718), (1450, 1672, 1660, 1718),
    (370, 1900, 570, 1936), (900, 1900, 1110, 1936), (1450, 1900, 1660, 1936),
    (370, 2120, 570, 2160), (900, 2120, 1110, 2160), (1450, 2120, 1660, 2160)
    ] 

    for i in range(30):
        voter_id_box = coordinates1[i]
        voter = im.crop(voter_id_box)
        text1 = ocr_tool.image_to_string( voter, builder=pyocr.builders.TextBuilder())
        text2 = text1.upper().strip()
        # if len(text2) > 10:
        #     text2 = text2[1:]
        text2=text2.replace("_",'').replace(']','').replace(" ",'').replace('|','').replace('-','').replace('[','').replace(',','').replace('—','').replace('™','TM').replace('O','0').replace('E','').replace('S','J')
        crop_rectangle_nm = coordinates[i]
        cropped_nm = im.crop(crop_rectangle_nm)
        text = pytesseract.image_to_string(cropped_nm)
        text = text.replace('=', '').replace('!', '').replace("'", '').replace('\n', ' ').replace(':', '').replace(
            '?', '').replace('$', '').replace('@', '').replace('#', '').replace('%', '').replace('*','').replace('¢','').replace(
            '+', '').replace('<', '').replace('>', '').replace(';', '').replace('©', '').replace('«', '')

        if text :
            patterns = {
                'Name': r'(?:Name|Nam)\s*\s*(.*?)\s*(?:Husbands|Fathers|Mothers|Others)',
                'Husbands Name': r'(?:Husbands Name|Fathers Name|Mothers Name|Others) \s*\s*(.*?)\s*House Number',
                'House Number': r'House Number\s*\s*(.*?)\s*(?:Age|Aga)',
                'Age': r'(?:Age|Aga)\s+\d+\s+(\d{2})\b\s*Gender',
                'Gender': r'Gender\s+\d*\s*(\w+)'
            }
            data = {}
            for attribute, pattern in patterns.items():
                match = re.search(pattern, text)
                if match:
                    data[attribute] = match.group(1).strip()
                else:
                    regex = r'(?:Age|Aga)\s+\d*(\d{2})\b\s*Gender'
                    matches = re.search(regex, text)
                    if matches:
                        data[attribute] = matches.group(1)
                if attribute == 'Gender':
                    data[attribute] = correct_gender_spelling(
                        match.group(1).strip()) if match else ''

            ws.append([ text2, data.get('Name', ''), data.get('Husbands Name', ''),
                       data.get('House Number', ''), data.get('Age', ''), data.get('Gender', '')])

def process_pdf_to_excel(pdf_path):
    # Convert PDF to images
    images = convert_from_path(pdf_path, poppler_path=r'C:\Users\anush\Downloads\Release-23.11.0-0\poppler-23.11.0\Library\bin')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Voter_ID','Name', 'Husbands Name/Fathers Name/Mothers Name', 'House Number', 'Age', 'Gender'])

    with ThreadPoolExecutor(max_workers=10) as executor:  # Adjust max_workers based on your system
        futures = []
        for idx, im in enumerate(images[2:-1], start=3):
            futures.append(executor.submit(process_image, im, idx, ws))

        for future in futures:
            future.result()
    ws.cell(row=1, column=ws.max_column + 1, value='DELETED')
    for row in ws.rows:
        count = 0
        for cell in row:
            if cell.value is None or cell.value == "":
                count += 1
        if count > 2:
            # print(row[0].row)
            ws.cell(row=row[0].row, column=ws.max_column, value='DELETED')
            
    output_excel_path = os.path.splitext(pdf_path)[0] + '.xlsx'
    wb.save(output_excel_path)
    return output_excel_path

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return render_template('index.html', error='No file part')

    file = request.files['file']

    if file.filename == '':
        return render_template('index.html', error='No selected file')

    if file:
        file_path = os.path.join('uploads', file.filename)
        file.save(file_path)

        output_excel_path = process_pdf_to_excel(file_path)

        return send_file(output_excel_path, as_attachment=True)

if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')

    app.run(debug=True)
